"""
Vista específica para la gestión de cultivos
"""
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
from views.base_view import BaseView
from utils.helpers import UIHelpers, safe_str
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.exceptions import EntityNotFoundError, EntityInUseError


class CultivosView(BaseView):

    def __init__(self, parent_frame, controller):
        self.form_title = "GESTIÓN DE CULTIVOS"
        self.entity_name = "CULTIVOS"
        self.imagen_miniatura = None
        self.label_imagen = None
        self.dark_theme = True
        super().__init__(parent_frame, controller)

    def _create_form_fields(self):


        self.create_form_field(0, "ID_CULTIVO:", "id_cultivo")


        self.create_form_field(1, "NOMBRE_CIENTIFICO:", "nombre_cientifico")


        self.create_form_field(2, "NOMBRE_COMUN:", "nombre_comun")


        self.create_form_field(3, "TIEMPO_CRECIMIENTO (días):", "tiempo_crecimiento")

        self.create_form_field(4, "TEMPERATURAS_OPTIMAS:", "temperaturas_optimas")

        self.create_form_field(5, "REQUERIMIENTO_AGUA:", "requerimiento_agua")

        self._create_photo_field()

    def _create_photo_field(self):


        photo_label = tk.Label(
            self.form_frame,
            text="PHOTO:",
            font=("Arial", 12)
        )
        photo_label.grid(row=6, column=0, sticky="w", padx=(0, 10), pady=8)

        self.photo_entry = tk.Entry(
            self.form_frame,
            width=25,
            font=("Arial", 12),
            relief="solid",
            bd=1
        )
        self.photo_entry.grid(row=6, column=1, sticky="w", pady=8)


        self.select_photo_btn = tk.Button(
            self.form_frame,
            text="Seleccionar Foto",
            font=("Arial", 10),
            bg="#2196F3",
            fg="white",
            command=self._on_select_photo
        )
        self.select_photo_btn.grid(row=6, column=2, padx=5, pady=8, sticky="w")

        self.label_imagen = tk.Label(
            self.form_frame,
            text="Sin imagen",
            font=("Arial", 10),
            fg="gray"
        )
        self.label_imagen.grid(row=7, column=1, columnspan=2, pady=10)


        self.form_fields['photo'] = self.photo_entry

    def _create_buttons(self):

        main_button_frame = tk.Frame(self.button_frame)
        main_button_frame.pack(pady=10)

        btn_save = tk.Button(
            main_button_frame,
            text="Guardar",
            font=("Arial", 10, "bold"),
            bg="#4CAF50",
            fg="white",
            width=10,
            command=self._on_save
        )
        btn_save.pack(side=tk.LEFT, padx=3)


        btn_update = tk.Button(
            main_button_frame,
            text="Actualizar",
            font=("Arial", 10, "bold"),
            bg="#2196F3",
            fg="white",
            width=10,
            command=self._on_update
        )
        btn_update.pack(side=tk.LEFT, padx=3)

        btn_delete = tk.Button(
            main_button_frame,
            text="Eliminar",
            font=("Arial", 10, "bold"),
            bg="#f44336",
            fg="white",
            width=10,
            command=self._on_delete
        )
        btn_delete.pack(side=tk.LEFT, padx=3)


        btn_clear = tk.Button(
            main_button_frame,
            text="Limpiar",
            font=("Arial", 10, "bold"),
            bg="#FF9800",
            fg="white",
            width=10,
            command=self._clear_form
        )
        btn_clear.pack(side=tk.LEFT, padx=3)

        search_frame = tk.Frame(self.button_frame)
        search_frame.pack(pady=5)

        btn_search = tk.Button(
            search_frame,
            text="Buscar por ID",
            font=("Arial", 10),
            bg="#9C27B0",
            fg="white",
            width=12,
            command=self._on_search
        )
        btn_search.pack(side=tk.LEFT, padx=3)

        self._create_additional_buttons()

    def _create_additional_buttons(self):

        additional_button_frame = tk.Frame(self.button_frame)
        additional_button_frame.pack(pady=10)

        btn_export_excel = tk.Button(
            additional_button_frame,
            text="Exportar Excel",
            font=("Arial", 10),
            bg="#4CAF50",
            fg="white",
            width=12,
            command=self._export_excel
        )
        btn_export_excel.pack(side=tk.LEFT, padx=3)

        btn_export_pdf = tk.Button(
            additional_button_frame,
            text="Exportar PDF",
            font=("Arial", 10),
            bg="#2196F3",
            fg="white",
            width=12,
            command=self._export_pdf
        )
        btn_export_pdf.pack(side=tk.LEFT, padx=3)

        btn_change_theme = tk.Button(
            additional_button_frame,
            text="Cambiar Tema",
            font=("Arial", 10),
            bg="#9E9E9E",
            fg="white",
            width=12,
            command=self._change_theme
        )
        btn_change_theme.pack(side=tk.LEFT, padx=3)

    def _create_treeview(self, parent):

        tree_frame = tk.Frame(parent)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = ('ID_CULTIVO', 'NOMBRE_CIENTIFICO', 'NOMBRE_COMUN',
                   'TIEMPO_CRECIMIENTO_DIAS', 'TEMPERATURAS_OPTIMAS',
                   'REQUERIMIENTO_AGUA_SEMANAL', 'PHOTO')

        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            height=20
        )


        self.tree.heading('ID_CULTIVO', text='ID')
        self.tree.heading('NOMBRE_CIENTIFICO', text='NOMBRE_CIENTIFICO')
        self.tree.heading('NOMBRE_COMUN', text='NOMBRE')
        self.tree.heading('TIEMPO_CRECIMIENTO_DIAS', text='TIEMPO_CRECIMIENTO')
        self.tree.heading('TEMPERATURAS_OPTIMAS', text='TEMP_OPTIMA')
        self.tree.heading('REQUERIMIENTO_AGUA_SEMANAL', text='REQUERIMIENTO_AGUA')
        self.tree.heading('PHOTO', text='PHOTO')

        self.tree.column('ID_CULTIVO', width=50, anchor='center')
        self.tree.column('NOMBRE_CIENTIFICO', width=90, anchor='w')
        self.tree.column('NOMBRE_COMUN', width=80, anchor='w')
        self.tree.column('TIEMPO_CRECIMIENTO_DIAS', width=80, anchor='center')
        self.tree.column('TEMPERATURAS_OPTIMAS', width=100, anchor='center')
        self.tree.column('REQUERIMIENTO_AGUA_SEMANAL', width=80, anchor='center')
        self.tree.column('PHOTO', width=70, anchor='center')


        self._configure_treeview_styles()


        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)


        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")


        self.tree.bind('<<TreeviewSelect>>', self._on_tree_select)

    def _on_tree_select(self, event):

        try:
            selected_items = self.tree.selection()
            if not selected_items:
                return

            selected_item = selected_items[0]
            values = self.tree.item(selected_item, 'values')

            if values and len(values) >= 7:
                data = {
                    'ID_CULTIVO': values[0],
                    'NOMBRE_CIENTIFICO': values[1],
                    'NOMBRE_COMUN': values[2],
                    'TIEMPO_CRECIMIENTO_DIAS': values[3],
                    'TEMPERATURAS_OPTIMAS': values[4],
                    'REQUERIMIENTO_AGUA_SEMANAL': values[5],
                    'PHOTO': values[6]
                }
                self._populate_form(data)

        except Exception as e:
            print(f"Error en selección de treeview: {e}")

    def _configure_treeview_styles(self, bg_color=None, fg_color=None, heading_bg=None, odd_bg=None, even_bg=None):

        try:
            style = ttk.Style()
            style.theme_use("default")


            if bg_color is None:
                bg_color = "#2b2b2b" if self.dark_theme else "white"
            if fg_color is None:
                fg_color = "white" if self.dark_theme else "black"
            if heading_bg is None:
                heading_bg = "#2ecc71" if self.dark_theme else "#4CAF50"
            if odd_bg is None:
                odd_bg = "#1e1e1e" if self.dark_theme else "#f9f9f9"
            if even_bg is None:
                even_bg = "#2e2e2e" if self.dark_theme else "#e9e9e9"


            style.configure("Treeview",
                            background=bg_color,
                            foreground=fg_color,
                            rowheight=25,
                            fieldbackground=bg_color,
                            font=('Arial', 11))

            style.configure("Treeview.Heading",
                            background=heading_bg,
                            foreground="white",
                            font=('Arial', 12, 'bold'))

            self.tree.tag_configure('oddrow', background=odd_bg)
            self.tree.tag_configure('evenrow', background=even_bg)

        except Exception as e:
            print(f"Error configurando estilos: {e}")

    def _get_form_data(self):

        form_data = {
            'ID_CULTIVO': self.get_field_value('id_cultivo'),
            'NOMBRE_CIENTIFICO': self.get_field_value('nombre_cientifico'),
            'NOMBRE_COMUN': self.get_field_value('nombre_comun'),
            'TIEMPO_CRECIMIENTO_DIAS': self.get_field_value('tiempo_crecimiento'),
            'TEMPERATURAS_OPTIMAS': self.get_field_value('temperaturas_optimas'),
            'REQUERIMIENTO_AGUA_SEMANAL': self.get_field_value('requerimiento_agua'),
            'PHOTO': self.get_field_value('photo')
        }
        print(f"Datos obtenidos del formulario: {form_data}")
        return form_data

    def _populate_form(self, data):

        try:
            if not data:
                return

            print(f"Poblando formulario con datos: {data}")

            self._clear_form()

            self.set_field_value('id_cultivo', data.get('ID_CULTIVO', ''))
            self.set_field_value('nombre_cientifico', data.get('NOMBRE_CIENTIFICO', ''))
            self.set_field_value('nombre_comun', data.get('NOMBRE_COMUN', ''))
            self.set_field_value('tiempo_crecimiento', data.get('TIEMPO_CRECIMIENTO_DIAS', ''))
            self.set_field_value('temperaturas_optimas', data.get('TEMPERATURAS_OPTIMAS', ''))
            self.set_field_value('requerimiento_agua', data.get('REQUERIMIENTO_AGUA_SEMANAL', ''))
            self.set_field_value('photo', data.get('PHOTO', ''))

            if data.get('PHOTO'):
                self.label_imagen.config(text=f"Imagen: {data.get('PHOTO', '').split('/')[-1]}")
            else:
                self.label_imagen.config(text="Sin imagen")

            print("Formulario poblado exitosamente")

        except Exception as e:
            print(f"Error poblando formulario: {e}")

    def _get_entity_id_from_form(self):

        cultivo_id = self.get_field_value('id_cultivo')
        try:
            return int(cultivo_id) if cultivo_id and cultivo_id.strip() else None
        except ValueError:
            return None

    def _refresh_list(self):

        try:

            for item in self.tree.get_children():
                self.tree.delete(item)

            cultivos = self.controller.get_all()
            print(f"Refrescando lista con {len(cultivos)} cultivos")


            for index, cultivo in enumerate(cultivos):
                values = (
                    safe_str(cultivo.get('ID_CULTIVO', '')),
                    safe_str(cultivo.get('NOMBRE_CIENTIFICO', '')),
                    safe_str(cultivo.get('NOMBRE_COMUN', '')),
                    safe_str(cultivo.get('TIEMPO_CRECIMIENTO_DIAS', '')),
                    safe_str(cultivo.get('TEMPERATURAS_OPTIMAS', '')),
                    safe_str(cultivo.get('REQUERIMIENTO_AGUA_SEMANAL', '')),
                    safe_str(cultivo.get('PHOTO', ''))
                )
                tag = 'evenrow' if index % 2 == 0 else 'oddrow'
                self.tree.insert('', 'end', values=values, tags=(tag,))

            print("Lista refrescada exitosamente")

        except Exception as e:
            print(f"Error en _refresh_list: {e}")
            import traceback
            traceback.print_exc()

    def _on_save(self):
        """Maneja el guardado de cultivos"""
        try:
            form_data = self._get_form_data()
            print("=== INICIANDO GUARDADO ===")
            print("Datos del formulario (GUARDAR):", form_data)

            # Validar campos obligatorios
            if not form_data.get('NOMBRE_CIENTIFICO') or not form_data.get('NOMBRE_CIENTIFICO').strip():
                messagebox.showerror("Error", "Debe ingresar el nombre científico")
                return

            if not form_data.get('NOMBRE_COMUN') or not form_data.get('NOMBRE_COMUN').strip():
                messagebox.showerror("Error", "Debe ingresar el nombre común")
                return

            if not form_data.get('TIEMPO_CRECIMIENTO_DIAS') or not form_data.get('TIEMPO_CRECIMIENTO_DIAS').strip():
                messagebox.showerror("Error", "Debe ingresar el tiempo de crecimiento (días)")
                return

            if not form_data.get('TEMPERATURAS_OPTIMAS') or not form_data.get('TEMPERATURAS_OPTIMAS').strip():
                messagebox.showerror("Error", "Debe ingresar la temperatura óptima")
                return

            # Crear nuevo cultivo
            print("Llamando a controller.create...")  # DEBUG
            new_id = self.controller.create(form_data)

            if new_id:
                messagebox.showinfo("Éxito", f"Cultivo guardado correctamente con ID: {new_id}")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "Error al guardar cultivo - No se retornó ID")

        except ValueError as ve:
            messagebox.showerror("Error de validación", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar cultivo: {str(e)}")
            print(f"ERROR DETALLADO en _on_save: {e}")  # DEBUG
            import traceback
            traceback.print_exc()

    def _on_update(self):
        """Maneja la actualización de cultivos"""
        try:
            cultivo_id = self._get_entity_id_from_form()
            print("=== INICIANDO ACTUALIZACIÓN ===")  # DEBUG
            print(f"ID para actualizar: {cultivo_id}")

            if not cultivo_id:
                messagebox.showerror("Error", "Seleccione un cultivo para actualizar")
                return

            form_data = self._get_form_data()
            print("Datos del formulario (ACTUALIZAR):", form_data)

            # Validar campos obligatorios
            if not form_data.get('NOMBRE_CIENTIFICO') or not form_data.get('NOMBRE_CIENTIFICO').strip():
                messagebox.showerror("Error", "Debe ingresar el nombre científico")
                return

            if not form_data.get('NOMBRE_COMUN') or not form_data.get('NOMBRE_COMUN').strip():
                messagebox.showerror("Error", "Debe ingresar el nombre común")
                return

            if not form_data.get('TIEMPO_CRECIMIENTO_DIAS') or not form_data.get('TIEMPO_CRECIMIENTO_DIAS').strip():
                messagebox.showerror("Error", "Debe ingresar el tiempo de crecimiento (días)")
                return

            if not form_data.get('TEMPERATURAS_OPTIMAS') or not form_data.get('TEMPERATURAS_OPTIMAS').strip():
                messagebox.showerror("Error", "Debe ingresar la temperatura óptima")
                return

            # Confirmación
            confirmar = messagebox.askyesno("Confirmación", f"¿Desea actualizar el cultivo con ID {cultivo_id}?")
            if not confirmar:
                return

            print("Llamando a controller.update...")  # DEBUG
            # Actualizar cultivo
            success = self.controller.update(cultivo_id, form_data)

            if success:
                messagebox.showinfo("Éxito", f"Cultivo con ID {cultivo_id} actualizado correctamente")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "No se pudo actualizar el cultivo")

        except ValueError as ve:
            messagebox.showerror("Error de validación", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Error actualizando cultivo: {str(e)}")
            print(f"ERROR DETALLADO en _on_update: {e}")  # DEBUG
            import traceback
            traceback.print_exc()

    def _on_delete(self):
        """Elimina un cultivo"""
        try:
            cultivo_id = self._get_entity_id_from_form()
            print("=== INICIANDO ELIMINACIÓN ===")  # DEBUG
            print(f"ID para eliminar: {cultivo_id}")

            if not cultivo_id:
                messagebox.showwarning("Advertencia", "Por favor, ingresa un ID de cultivo válido")
                return

            # Confirmación
            respuesta = messagebox.askyesno("Confirmar eliminación",
                                            f"¿Estás seguro de eliminar el cultivo con ID {cultivo_id}?")
            if not respuesta:
                messagebox.showinfo("Cancelado", "Eliminación cancelada")
                return

            print("Llamando a controller.delete...")  # DEBUG
            success = self.controller.delete(cultivo_id)

            if success:
                messagebox.showinfo("Éxito", f"Cultivo con ID {cultivo_id} eliminado correctamente")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "No se pudo eliminar el cultivo")

        except EntityNotFoundError as e:
            messagebox.showerror("No encontrado", f"No se encontró el cultivo con ID {cultivo_id}")
            self._clear_form()
            self._refresh_list()
        except EntityInUseError as e:
            messagebox.showerror(
                "No se puede eliminar",
                f"No se puede eliminar el cultivo porque {e.reason}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar el cultivo: {str(e)}")
            print(f"ERROR DETALLADO en _on_delete: {e}")  # DEBUG
            import traceback
            traceback.print_exc()

    def _on_search(self):
        """Busca un cultivo por ID"""
        cultivo_id = self.get_field_value('id_cultivo')

        try:
            print("=== INICIANDO BÚSQUEDA ===")  # DEBUG
            print(f"ID a buscar: {cultivo_id}")

            if not cultivo_id or not cultivo_id.strip():
                messagebox.showerror("Error", "Ingrese un ID de cultivo para buscar")
                return

            # Validar que sea un número válido
            entity_id = int(cultivo_id.strip())

            if entity_id <= 0:
                messagebox.showerror("Error", "El ID del cultivo debe ser un número positivo")
                return

            print("Llamando a controller.get_by_id...")  # DEBUG
            # Llamar al controlador para buscar
            data = self.controller.get_by_id(entity_id)

            if data:
                self._populate_form(data)
                messagebox.showinfo("Éxito", "Cultivo encontrado y cargado en el formulario")
            else:
                messagebox.showerror("No encontrado", f"No se encontró ningún cultivo con ID: {entity_id}")

        except ValueError:
            messagebox.showerror("Error", "El ID del cultivo debe ser un número válido")
        except Exception as e:
            messagebox.showerror("Error", f"Error en búsqueda: {str(e)}")
            print(f"ERROR DETALLADO en _on_search: {e}")  # DEBUG

    def _on_select_photo(self):
        """Maneja la selección de foto"""
        try:
            file_path = filedialog.askopenfilename(
                title="Seleccionar imagen",
                filetypes=[("Image files", "*.jpg *.jpeg *.png *.gif *.bmp")]
            )

            if file_path:
                self.photo_entry.delete(0, tk.END)
                self.photo_entry.insert(0, file_path)
                self.label_imagen.config(text=f"Imagen: {file_path.split('/')[-1]}")

        except Exception as e:
            messagebox.showerror("Error", f"Error seleccionando foto: {str(e)}")

    def _change_theme(self):
        """Cambia el tema de la aplicación"""
        try:
            self.dark_theme = not self.dark_theme
            self._apply_theme()
            messagebox.showinfo("Tema Cambiado", f"Tema {'oscuro' if self.dark_theme else 'claro'} aplicado")
        except Exception as e:
            messagebox.showerror("Error", f"Error cambiando tema: {str(e)}")

    def _apply_theme(self):
        """Aplica el tema actual a todos los elementos de la vista"""
        try:
            if self.dark_theme:
                # Tema oscuro
                bg_color = "#2b2b2b"
                fg_color = "white"
                entry_bg = "#1e1e1e"
                entry_fg = "white"
                button_bg = "#2196F3"
                tree_bg = "#2b2b2b"
                tree_fg = "white"
                heading_bg = "#2ecc71"
                odd_bg = "#1e1e1e"
                even_bg = "#2e2e2e"
            else:
                # Tema claro
                bg_color = "#f0f0f0"
                fg_color = "black"
                entry_bg = "white"
                entry_fg = "black"
                button_bg = "#2196F3"
                tree_bg = "white"
                tree_fg = "black"
                heading_bg = "#4CAF50"
                odd_bg = "#f9f9f9"
                even_bg = "#e9e9e9"

            # Aplicar colores a frames principales
            self.main_frame.configure(bg=bg_color)
            self.form_frame.configure(bg=bg_color)
            self.button_frame.configure(bg=bg_color)

            # Aplicar tema recursivamente
            self._apply_theme_to_container(self.main_frame, bg_color, fg_color, entry_bg, entry_fg, button_bg)
            self._apply_theme_to_container(self.form_frame, bg_color, fg_color, entry_bg, entry_fg, button_bg)
            self._apply_theme_to_container(self.button_frame, bg_color, fg_color, entry_bg, entry_fg, button_bg)

            # Aplicar tema al TreeView
            self._configure_treeview_styles(tree_bg, tree_fg, heading_bg, odd_bg, even_bg)

        except Exception as e:
            print(f"Error aplicando tema: {e}")

    def _apply_theme_to_container(self, container, bg_color, fg_color, entry_bg, entry_fg, button_bg):
        """Aplica el tema recursivamente a un contenedor y todos sus hijos"""
        try:
            # Aplicar al contenedor mismo
            if isinstance(container, (tk.Frame, tk.LabelFrame)):
                container.configure(bg=bg_color)

            # Aplicar a todos los hijos del contenedor
            for widget in container.winfo_children():
                self._apply_theme_to_widget(widget, bg_color, fg_color, entry_bg, entry_fg, button_bg)

                # Si el widget es un contenedor, aplicar recursivamente a sus hijos
                if isinstance(widget, (tk.Frame, tk.LabelFrame)):
                    self._apply_theme_to_container(widget, bg_color, fg_color, entry_bg, entry_fg, button_bg)

        except Exception as e:
            print(f"Error aplicando tema al contenedor: {e}")

    def _apply_theme_to_widget(self, widget, bg_color, fg_color, entry_bg, entry_fg, button_bg):
        """Aplica el tema a un widget específico"""
        try:
            if isinstance(widget, tk.Label):
                widget.configure(bg=bg_color, fg=fg_color)
            elif isinstance(widget, tk.Entry):
                widget.configure(bg=entry_bg, fg=entry_fg, insertbackground=fg_color)
            elif isinstance(widget, tk.Button):
                current_text = widget.cget('text')
                # Mantener colores específicos para botones de acción
                if current_text in ['Guardar', 'Actualizar', 'Eliminar', 'Limpiar', 'Buscar']:
                    pass
                elif current_text == 'Cambiar Tema':
                    widget.configure(bg="#9E9E9E", fg="white")
                elif current_text == 'Exportar Excel':
                    widget.configure(bg="#4CAF50", fg="white")
                elif current_text == 'Exportar PDF':
                    widget.configure(bg="#2196F3", fg="white")
                elif current_text == 'Seleccionar Foto':
                    widget.configure(bg="#2196F3", fg="white")
                else:
                    widget.configure(bg=button_bg, fg="white")
            elif isinstance(widget, tk.Frame):
                widget.configure(bg=bg_color)
        except Exception as e:
            pass

    def _export_excel(self):
        """Exporta datos de cultivos a Excel con formato profesional"""
        try:
            cultivos = self.controller.get_all()
            if not cultivos:
                messagebox.showerror("Error", "No hay datos para exportar")
                return

            # Pedir al usuario donde guardar el Excel
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar Excel como"
            )

            if not file_path:
                return  # Usuario canceló

            # Crear libro de trabajo y hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Cultivos"

            # Estilos
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
            normal_font = Font(name='Arial', size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')

            # Título
            ws.merge_cells('A1:G1')
            ws['A1'] = 'REPORTE DE CULTIVOS'
            ws['A1'].font = Font(name='Arial', size=14, bold=True)
            ws['A1'].alignment = center_align

            # Fecha
            ws.merge_cells('A2:G2')
            ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
            ws['A2'].font = normal_font
            ws['A2'].alignment = center_align

            # Espacio
            ws.append([])

            # Encabezados de columnas
            headers = [
                'ID CULTIVO',
                'NOMBRE CIENTÍFICO',
                'NOMBRE COMÚN',
                'TIEMPO CRECIMIENTO (días)',
                'TEMPERATURAS ÓPTIMAS',
                'REQUERIMIENTO AGUA SEMANAL',
                'RUTA FOTO'
            ]
            ws.append(headers)

            # Aplicar estilo a los encabezados
            for col in range(1, 8):  # Columnas A a G
                cell = ws.cell(row=4, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            # Datos de los cultivos
            for cultivo in cultivos:
                row = [
                    safe_str(cultivo.get('ID_CULTIVO', '')),
                    safe_str(cultivo.get('NOMBRE_CIENTIFICO', '')),
                    safe_str(cultivo.get('NOMBRE_COMUN', '')),
                    safe_str(cultivo.get('TIEMPO_CRECIMIENTO_DIAS', '')),
                    safe_str(cultivo.get('TEMPERATURAS_OPTIMAS', '')),
                    safe_str(cultivo.get('REQUERIMIENTO_AGUA_SEMANAL', '')),
                    safe_str(cultivo.get('PHOTO', ''))
                ]
                ws.append(row)

            # Aplicar estilo a los datos
            for row in range(5, len(cultivos) + 5):
                # Alternar colores de fila
                if row % 2 == 1:
                    fill_color = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
                else:
                    fill_color = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = border
                    cell.fill = fill_color

                    # Alineación específica por columna
                    if col in [1, 4]:  # ID y Tiempo Crecimiento - centrados
                        cell.alignment = center_align
                    else:  # Texto - alineado a la izquierda
                        cell.alignment = left_align

            # Ajustar anchos de columna
            column_widths = {
                'A': 10,  # ID
                'B': 25,  # Nombre Científico
                'C': 20,  # Nombre Común
                'D': 18,  # Tiempo Crecimiento
                'E': 20,  # Temperaturas
                'F': 22,  # Requerimiento Agua
                'G': 30  # Ruta Foto
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Resumen
            summary_row = len(cultivos) + 6
            ws.merge_cells(f'A{summary_row}:G{summary_row}')
            ws[f'A{summary_row}'] = f'Total de cultivos registrados: {len(cultivos)}'
            ws[f'A{summary_row}'].font = Font(name='Arial', size=11, bold=True)
            ws[f'A{summary_row}'].alignment = center_align
            ws[f'A{summary_row}'].fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

            # Guardar el archivo
            wb.save(file_path)

            messagebox.showinfo("Éxito", f"Excel exportado correctamente:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exportando a Excel: {str(e)}")

    def _export_pdf(self):
        """Exporta datos de cultivos a PDF"""
        try:
            cultivos = self.controller.get_all()
            if not cultivos:
                messagebox.showerror("Error", "No hay datos para exportar")
                return

            # Pedir al usuario donde guardar el PDF
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Guardar PDF como"
            )

            if not file_path:
                return  # Usuario canceló

            # Crear el documento PDF
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )

            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1,  # Centrado
                textColor=colors.darkblue
            )

            # Contenido del PDF
            elements = []

            # Título
            title = Paragraph("REPORTE DE CULTIVOS", title_style)
            elements.append(title)

            # Fecha de generación
            fecha = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal'])
            elements.append(fecha)
            elements.append(Spacer(1, 20))

            # Preparar datos para la tabla
            table_data = [
                ['ID', 'Nombre Científico', 'Nombre Común', 'Tiempo Crecimiento', 'Temperaturas', 'Agua Semanal']]

            for cultivo in cultivos:
                table_data.append([
                    safe_str(cultivo.get('ID_CULTIVO', '')),
                    safe_str(cultivo.get('NOMBRE_CIENTIFICO', '')),
                    safe_str(cultivo.get('NOMBRE_COMUN', '')),
                    safe_str(cultivo.get('TIEMPO_CRECIMIENTO_DIAS', '')),
                    safe_str(cultivo.get('TEMPERATURAS_OPTIMAS', '')),
                    safe_str(cultivo.get('REQUERIMIENTO_AGUA_SEMANAL', ''))
                ])

            # Crear tabla
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                # Estilo del encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),

                # Estilo de las filas
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),

                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            # Alternar colores de filas
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    bg_color = colors.HexColor('#f8f8f8')
                else:
                    bg_color = colors.white
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), bg_color)
                ]))

            elements.append(table)
            elements.append(Spacer(1, 20))

            # Resumen
            total_cultivos = len(cultivos)
            resumen = Paragraph(f"Total de cultivos registrados: {total_cultivos}", styles['Normal'])
            elements.append(resumen)

            # Generar PDF
            doc.build(elements)

            messagebox.showinfo("Éxito", f"PDF exportado correctamente:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exportando a PDF: {str(e)}")

    def _clear_form(self):
        """Limpia el formulario y resetea validaciones"""
        super()._clear_form()

        # Limpiar campo de foto
        if hasattr(self, 'label_imagen'):
            self.label_imagen.config(text="Sin imagen")

        # Enfocar el primer campo
        self.focus_field('id_cultivo')
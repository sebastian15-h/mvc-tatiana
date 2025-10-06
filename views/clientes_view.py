"""
Vista específica para la gestión de clientes
"""
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
from views.base_view import BaseView
from utils.helpers import safe_str
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.exceptions import EntityNotFoundError, EntityInUseError


class ClientesView(BaseView):
    """Vista específica para clientes"""

    def __init__(self, parent_frame, controller):
        self.form_title = "GESTIÓN DE CLIENTES"
        self.entity_name = "CLIENTES"
        self.dark_theme = True
        super().__init__(parent_frame, controller)

    def _create_form_fields(self):
        """Crea los campos específicos del formulario de clientes"""
        # Campo ID_CLIENTE
        self.create_form_field(0, "ID_CLIENTE:", "id_cliente")

        # Campo NOMBRE
        self.create_form_field(1, "NOMBRE:", "nombre")

        # Campo APELLIDO
        self.create_form_field(2, "APELLIDO:", "apellido")

        # Campo DOCUMENTO_IDENTIDAD
        self.create_form_field(3, "DOCUMENTO_IDENTIDAD:", "documento_identidad")

        # Campo NACIONALIDAD
        self.create_form_field(4, "NACIONALIDAD:", "nacionalidad")

        # Campo FECHA_NACIMIENTO
        self.create_form_field(5, "FECHA_NACIMIENTO:", "fecha_nacimiento")

        # Campo DIRECCION
        self.create_form_field(6, "DIRECCION:", "direccion")

        # Campo TELEFONO
        self.create_form_field(7, "TELEFONO:", "telefono")

        # Campo CORREO
        self.create_form_field(8, "CORREO:", "correo")

        # Campo PREFERENCIAS_ESPECIALES
        self.create_form_field(9, "PREFERENCIAS_ESPECIALES:", "preferencias_especiales")

        # Campo NIVEL_PROGRAMA_FIDELIZACION
        self.create_form_field(10, "NIVEL_PROGRAMA_FIDELIZACION:", "nivel_programa_fidelizacion")

    def _create_buttons(self):
        """Crea los botones de acción - Sobrescribe el método de BaseView"""
        # Frame para botones principales
        main_button_frame = tk.Frame(self.button_frame)
        main_button_frame.pack(pady=10)

        # Botón Guardar (Crear nuevo)
        btn_save = tk.Button(
            main_button_frame,
            text="Guardar",
            font=("Arial", 10, "bold"),
            bg="#4CAF50",
            fg="white",
            width=10,
            command=self._on_save
        )
        btn_save.pack(side=tk.LEFT, padx=3)

        # Botón Actualizar
        btn_update = tk.Button(
            main_button_frame,
            text="Actualizar",
            font=("Arial", 10, "bold"),
            bg="#2196F3",
            fg="white",
            width=10,
            command=self._on_update
        )
        btn_update.pack(side=tk.LEFT, padx=3)

        # Botón Eliminar
        btn_delete = tk.Button(
            main_button_frame,
            text="Eliminar",
            font=("Arial", 10, "bold"),
            bg="#f44336",
            fg="white",
            width=10,
            command=self._on_delete
        )
        btn_delete.pack(side=tk.LEFT, padx=3)

        # Botón Limpiar
        btn_clear = tk.Button(
            main_button_frame,
            text="Limpiar",
            font=("Arial", 10, "bold"),
            bg="#FF9800",
            fg="white",
            width=10,
            command=self._clear_form
        )
        btn_clear.pack(side=tk.LEFT, padx=3)

        # Frame para búsqueda
        search_frame = tk.Frame(self.button_frame)
        search_frame.pack(pady=5)

        # Botón Buscar
        btn_search = tk.Button(
            search_frame,
            text="Buscar por ID",
            font=("Arial", 10),
            bg="#9C27B0",
            fg="white",
            width=12,
            command=self._on_search
        )
        btn_search.pack(side=tk.LEFT, padx=3)

        # Luego añadir los botones adicionales
        self._create_additional_buttons()

    def _create_additional_buttons(self):
        """Crea botones adicionales específicos para clientes"""
        # Frame para botones adicionales
        additional_button_frame = tk.Frame(self.button_frame)
        additional_button_frame.pack(pady=10)

        # Botón Exportar Excel
        btn_export_excel = tk.Button(
            additional_button_frame,
            text="Exportar Excel",
            font=("Arial", 10),
            bg="#4CAF50",
            fg="white",
            width=12,
            command=self._export_excel
        )
        btn_export_excel.pack(side=tk.LEFT, padx=3)

        # Botón Exportar PDF
        btn_export_pdf = tk.Button(
            additional_button_frame,
            text="Exportar PDF",
            font=("Arial", 10),
            bg="#2196F3",
            fg="white",
            width=12,
            command=self._export_pdf
        )
        btn_export_pdf.pack(side=tk.LEFT, padx=3)

        # Botón Cambiar Tema
        btn_change_theme = tk.Button(
            additional_button_frame,
            text="Cambiar Tema",
            font=("Arial", 10),
            bg="#9E9E9E",
            fg="white",
            width=12,
            command=self._change_theme
        )
        btn_change_theme.pack(side=tk.LEFT, padx=3)

    def _create_treeview(self, parent):
        """Crea el TreeView específico para clientes"""
        # Frame para TreeView y scrollbar
        tree_frame = tk.Frame(parent)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Definir columnas
        columns = ('ID_CLIENTE', 'NOMBRE', 'APELLIDO', 'DOCUMENTO_IDENTIDAD',
                   'NACIONALIDAD', 'FECHA_NACIMIENTO', 'DIRECCION', 'TELEFONO',
                   'CORREO', 'PREFERENCIAS_ESPECIALES', 'NIVEL_PROGRAMA_FIDELIZACION')

        # Crear TreeView
        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            height=20
        )

        # Configurar encabezados
        self.tree.heading('ID_CLIENTE', text='ID')
        self.tree.heading('NOMBRE', text='NOMBRE')
        self.tree.heading('APELLIDO', text='APELLIDO')
        self.tree.heading('DOCUMENTO_IDENTIDAD', text='DOCUMENTO')
        self.tree.heading('NACIONALIDAD', text='NACIONALIDAD')
        self.tree.heading('FECHA_NACIMIENTO', text='FECHA NACIMIENTO')
        self.tree.heading('DIRECCION', text='DIRECCIÓN')
        self.tree.heading('TELEFONO', text='TELÉFONO')
        self.tree.heading('CORREO', text='CORREO')
        self.tree.heading('PREFERENCIAS_ESPECIALES', text='PREFERENCIAS')
        self.tree.heading('NIVEL_PROGRAMA_FIDELIZACION', text='NIVEL FIDELIZACIÓN')

        # Configurar anchos de columnas
        self.tree.column('ID_CLIENTE', width=60, anchor='center')
        self.tree.column('NOMBRE', width=100, anchor='w')
        self.tree.column('APELLIDO', width=100, anchor='w')
        self.tree.column('DOCUMENTO_IDENTIDAD', width=100, anchor='center')
        self.tree.column('NACIONALIDAD', width=100, anchor='w')
        self.tree.column('FECHA_NACIMIENTO', width=100, anchor='center')
        self.tree.column('DIRECCION', width=120, anchor='w')
        self.tree.column('TELEFONO', width=100, anchor='center')
        self.tree.column('CORREO', width=120, anchor='w')
        self.tree.column('PREFERENCIAS_ESPECIALES', width=120, anchor='w')
        self.tree.column('NIVEL_PROGRAMA_FIDELIZACION', width=100, anchor='center')

        # Configurar estilos iniciales
        self._configure_treeview_styles()

        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Layout
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Evento de selección
        self.tree.bind('<<TreeviewSelect>>', self._on_tree_select)

    def _configure_treeview_styles(self):
        """Configura los estilos para el Treeview"""
        style = ttk.Style()

        # Configurar estilo para el Treeview
        style.configure("Treeview",
                        background="#ffffff",
                        foreground="#000000",
                        rowheight=25,
                        fieldbackground="#ffffff")

        style.configure("Treeview.Heading",
                        background="#4CAF50",
                        foreground="white",
                        relief="flat",
                        font=('Arial', 10, 'bold'))

        style.map("Treeview.Heading",
                  background=[('active', '#45a049')])

        style.map("Treeview",
                  background=[('selected', '#4CAF50')],
                  foreground=[('selected', 'white')])

        # Configurar estilos para filas alternadas
        style.configure("Treeview",
                        rowheight=25,
                        font=('Arial', 9))

        # Configurar tags para filas alternadas
        self.tree.tag_configure('evenrow', background='#f8f8f8')
        self.tree.tag_configure('oddrow', background='#ffffff')

    def _on_tree_select(self, event):
        """Maneja la selección de un item en el treeview"""
        try:
            selected_items = self.tree.selection()
            if not selected_items:
                return

            selected_item = selected_items[0]
            values = self.tree.item(selected_item, 'values')

            if values and len(values) >= 11:
                data = {
                    'ID_CLIENTE': values[0],
                    'NOMBRE': values[1],
                    'APELLIDO': values[2],
                    'DOCUMENTO_IDENTIDAD': values[3],
                    'NACIONALIDAD': values[4],
                    'FECHA_NACIMIENTO': values[5],
                    'DIRECCION': values[6],
                    'TELEFONO': values[7],
                    'CORREO': values[8],
                    'PREFERENCIAS_ESPECIALES': values[9],
                    'NIVEL_PROGRAMA_FIDELIZACION': values[10]
                }
                print(f"Datos seleccionados del treeview: {data}")
                self._populate_form(data)

        except Exception as e:
            print(f"Error en selección de treeview: {e}")

    def _get_form_data(self):
        """Obtiene los datos del formulario de clientes"""
        form_data = {
            'ID_CLIENTE': self.get_field_value('id_cliente'),
            'NOMBRE': self.get_field_value('nombre'),
            'APELLIDO': self.get_field_value('apellido'),
            'DOCUMENTO_IDENTIDAD': self.get_field_value('documento_identidad'),
            'NACIONALIDAD': self.get_field_value('nacionalidad'),
            'FECHA_NACIMIENTO': self.get_field_value('fecha_nacimiento'),
            'DIRECCION': self.get_field_value('direccion'),
            'TELEFONO': self.get_field_value('telefono'),
            'CORREO': self.get_field_value('correo'),
            'PREFERENCIAS_ESPECIALES': self.get_field_value('preferencias_especiales'),
            'NIVEL_PROGRAMA_FIDELIZACION': self.get_field_value('nivel_programa_fidelizacion')
        }
        print(f"Datos obtenidos del formulario: {form_data}")
        return form_data

    def _populate_form(self, data):
        """Puebla el formulario con datos de un cliente"""
        try:
            if not data:
                return

            print(f"Poblando formulario con datos: {data}")

            # Limpiar formulario primero
            self._clear_form()

            # Poblar campos con los datos
            self.set_field_value('id_cliente', data.get('ID_CLIENTE', ''))
            self.set_field_value('nombre', data.get('NOMBRE', ''))
            self.set_field_value('apellido', data.get('APELLIDO', ''))
            self.set_field_value('documento_identidad', data.get('DOCUMENTO_IDENTIDAD', ''))
            self.set_field_value('nacionalidad', data.get('NACIONALIDAD', ''))
            self.set_field_value('fecha_nacimiento', data.get('FECHA_NACIMIENTO', ''))
            self.set_field_value('direccion', data.get('DIRECCION', ''))
            self.set_field_value('telefono', data.get('TELEFONO', ''))
            self.set_field_value('correo', data.get('CORREO', ''))
            self.set_field_value('preferencias_especiales', data.get('PREFERENCIAS_ESPECIALES', ''))
            self.set_field_value('nivel_programa_fidelizacion', data.get('NIVEL_PROGRAMA_FIDELIZACION', ''))

            print("Formulario poblado exitosamente")

        except Exception as e:
            print(f"Error poblando formulario: {e}")

    def _get_entity_id_from_form(self):
        """Obtiene el ID del cliente desde el formulario"""
        cliente_id = self.get_field_value('id_cliente')
        try:
            return int(cliente_id) if cliente_id and cliente_id.strip() else None
        except ValueError:
            return None

    def _refresh_list(self):
        """Actualiza la lista de clientes con colores alternados"""
        try:
            # Limpiar TreeView completamente
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Obtener datos actualizados del controlador
            clientes = self.controller.get_all()
            print(f"Refrescando lista con {len(clientes)} clientes")

            # Insertar datos actualizados
            for index, cliente in enumerate(clientes):
                values = (
                    safe_str(cliente.get('ID_CLIENTE', '')),
                    safe_str(cliente.get('NOMBRE', '')),
                    safe_str(cliente.get('APELLIDO', '')),
                    safe_str(cliente.get('DOCUMENTO_IDENTIDAD', '')),
                    safe_str(cliente.get('NACIONALIDAD', '')),
                    safe_str(cliente.get('FECHA_NACIMIENTO', '')),
                    safe_str(cliente.get('DIRECCION', '')),
                    safe_str(cliente.get('TELEFONO', '')),
                    safe_str(cliente.get('CORREO', '')),
                    safe_str(cliente.get('PREFERENCIAS_ESPECIALES', '')),
                    safe_str(cliente.get('NIVEL_PROGRAMA_FIDELIZACION', ''))
                )
                tag = 'evenrow' if index % 2 == 0 else 'oddrow'
                self.tree.insert('', 'end', values=values, tags=(tag,))

            print("Lista refrescada exitosamente")

        except Exception as e:
            print(f"Error en _refresh_list: {e}")
            import traceback
            traceback.print_exc()

    def _on_save(self):
        """Maneja el guardado de clientes"""
        try:
            form_data = self._get_form_data()
            print("=== INICIANDO GUARDADO ===")
            print("Datos del formulario (GUARDAR):", form_data)

            # Validar campos obligatorios
            if not form_data.get('NOMBRE') or not form_data.get('NOMBRE').strip():
                messagebox.showerror("Error", "Debe ingresar el nombre del cliente")
                return

            if not form_data.get('APELLIDO') or not form_data.get('APELLIDO').strip():
                messagebox.showerror("Error", "Debe ingresar el apellido del cliente")
                return

            if not form_data.get('DOCUMENTO_IDENTIDAD') or not form_data.get('DOCUMENTO_IDENTIDAD').strip():
                messagebox.showerror("Error", "Debe ingresar el documento de identidad")
                return

            # Crear nuevo cliente
            print("Llamando a controller.create...")
            new_id = self.controller.create(form_data)

            if new_id:
                messagebox.showinfo("Éxito", f"Cliente guardado correctamente con ID: {new_id}")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "Error al guardar cliente - No se retornó ID")

        except ValueError as ve:
            messagebox.showerror("Error de validación", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar cliente: {str(e)}")
            print(f"ERROR DETALLADO en _on_save: {e}")
            import traceback
            traceback.print_exc()

    def _on_update(self):
        """Maneja la actualización de clientes"""
        try:
            cliente_id = self._get_entity_id_from_form()
            print("=== INICIANDO ACTUALIZACIÓN ===")
            print(f"ID para actualizar: {cliente_id}")

            if not cliente_id:
                messagebox.showerror("Error", "Seleccione un cliente para actualizar")
                return

            form_data = self._get_form_data()
            print("Datos del formulario (ACTUALIZAR):", form_data)

            # Validar campos obligatorios
            if not form_data.get('NOMBRE') or not form_data.get('NOMBRE').strip():
                messagebox.showerror("Error", "Debe ingresar el nombre del cliente")
                return

            if not form_data.get('APELLIDO') or not form_data.get('APELLIDO').strip():
                messagebox.showerror("Error", "Debe ingresar el apellido del cliente")
                return

            if not form_data.get('DOCUMENTO_IDENTIDAD') or not form_data.get('DOCUMENTO_IDENTIDAD').strip():
                messagebox.showerror("Error", "Debe ingresar el documento de identidad")
                return

            # Confirmación
            confirmar = messagebox.askyesno("Confirmación", f"¿Desea actualizar el cliente con ID {cliente_id}?")
            if not confirmar:
                return

            print("Llamando a controller.update...")
            # Actualizar cliente
            success = self.controller.update(cliente_id, form_data)

            if success:
                messagebox.showinfo("Éxito", f"Cliente con ID {cliente_id} actualizado correctamente")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "No se pudo actualizar el cliente")

        except ValueError as ve:
            messagebox.showerror("Error de validación", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Error actualizando cliente: {str(e)}")
            print(f"ERROR DETALLADO en _on_update: {e}")
            import traceback
            traceback.print_exc()

    def _on_delete(self):
        """Elimina un cliente"""
        try:
            cliente_id = self._get_entity_id_from_form()
            print("=== INICIANDO ELIMINACIÓN ===")
            print(f"ID para eliminar: {cliente_id}")

            if not cliente_id:
                messagebox.showwarning("Advertencia", "Seleccione un cliente para eliminar")
                return

            # Confirmación
            respuesta = messagebox.askyesno("Confirmar eliminación",
                                            f"¿Estás seguro de eliminar el cliente con ID {cliente_id}?")
            if not respuesta:
                messagebox.showinfo("Cancelado", "Eliminación cancelada")
                return

            print("Llamando a controller.delete...")
            success = self.controller.delete(cliente_id)

            if success:
                messagebox.showinfo("Éxito", f"Cliente con ID {cliente_id} eliminado correctamente")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "No se pudo eliminar el cliente")

        except EntityNotFoundError as e:
            messagebox.showerror("No encontrado", f"No se encontró el cliente con ID {cliente_id}")
            self._clear_form()
            self._refresh_list()
        except EntityInUseError as e:
            messagebox.showerror(
                "No se puede eliminar",
                f"No se puede eliminar el cliente porque {e.reason}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar el cliente: {str(e)}")
            print(f"ERROR DETALLADO en _on_delete: {e}")
            import traceback
            traceback.print_exc()

    def _on_search(self):
        """Busca un cliente por ID"""
        cliente_id = self.get_field_value('id_cliente')

        try:
            print("=== INICIANDO BÚSQUEDA ===")
            print(f"ID a buscar: {cliente_id}")

            if not cliente_id or not cliente_id.strip():
                messagebox.showerror("Error", "Ingrese un ID de cliente para buscar")
                return

            # Validar que sea un número válido
            entity_id = int(cliente_id.strip())

            if entity_id <= 0:
                messagebox.showerror("Error", "El ID del cliente debe ser un número positivo")
                return

            print("Llamando a controller.get_by_id...")
            # Llamar al controlador para buscar
            data = self.controller.get_by_id(entity_id)

            if data:
                self._populate_form(data)
                messagebox.showinfo("Éxito", "Cliente encontrado y cargado en el formulario")
            else:
                messagebox.showerror("No encontrado", f"No se encontró ningún cliente con ID: {entity_id}")

        except ValueError:
            messagebox.showerror("Error", "El ID del cliente debe ser un número válido")
        except Exception as e:
            messagebox.showerror("Error", f"Error en búsqueda: {str(e)}")
            print(f"ERROR DETALLADO en _on_search: {e}")

    def _export_excel(self):
        """Exporta datos de clientes a Excel"""
        try:
            clientes = self.controller.get_all()
            if not clientes:
                messagebox.showerror("Error", "No hay datos para exportar")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar Excel como",
                initialfile="clientes.xlsx"
            )

            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Clientes"

            # Estilos (similar a hoteles_view)
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
            normal_font = Font(name='Arial', size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')

            # Título
            ws.merge_cells('A1:K1')
            ws['A1'] = 'REPORTE DE CLIENTES'
            ws['A1'].font = Font(name='Arial', size=14, bold=True)
            ws['A1'].alignment = center_align

            # Fecha
            ws.merge_cells('A2:K2')
            ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
            ws['A2'].font = normal_font
            ws['A2'].alignment = center_align

            # Espacio
            ws.append([])

            # Encabezados
            headers = [
                'ID CLIENTE', 'NOMBRE', 'APELLIDO', 'DOCUMENTO IDENTIDAD',
                'NACIONALIDAD', 'FECHA NACIMIENTO', 'DIRECCIÓN', 'TELÉFONO',
                'CORREO', 'PREFERENCIAS ESPECIALES', 'NIVEL FIDELIZACIÓN'
            ]
            ws.append(headers)

            # Aplicar estilo a los encabezados
            for col in range(1, 12):
                cell = ws.cell(row=4, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            # Datos
            for cliente in clientes:
                row = [
                    safe_str(cliente.get('ID_CLIENTE', '')),
                    safe_str(cliente.get('NOMBRE', '')),
                    safe_str(cliente.get('APELLIDO', '')),
                    safe_str(cliente.get('DOCUMENTO_IDENTIDAD', '')),
                    safe_str(cliente.get('NACIONALIDAD', '')),
                    safe_str(cliente.get('FECHA_NACIMIENTO', '')),
                    safe_str(cliente.get('DIRECCION', '')),
                    safe_str(cliente.get('TELEFONO', '')),
                    safe_str(cliente.get('CORREO', '')),
                    safe_str(cliente.get('PREFERENCIAS_ESPECIALES', '')),
                    safe_str(cliente.get('NIVEL_PROGRAMA_FIDELIZACION', ''))
                ]
                ws.append(row)

            # Aplicar estilo a los datos
            for row in range(5, len(clientes) + 5):
                fill_color = PatternFill(
                    start_color='F0F8FF' if row % 2 == 1 else 'FFFFFF',
                    end_color='F0F8FF' if row % 2 == 1 else 'FFFFFF',
                    fill_type='solid'
                )

                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = border
                    cell.fill = fill_color
                    cell.alignment = left_align

            # Ajustar anchos
            column_widths = {
                'A': 10, 'B': 15, 'C': 15, 'D': 15, 'E': 12,
                'F': 15, 'G': 20, 'H': 12, 'I': 20, 'J': 20, 'K': 15
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Resumen
            summary_row = len(clientes) + 6
            ws.merge_cells(f'A{summary_row}:K{summary_row}')
            ws[f'A{summary_row}'] = f'Total de clientes registrados: {len(clientes)}'
            ws[f'A{summary_row}'].font = Font(name='Arial', size=11, bold=True)
            ws[f'A{summary_row}'].alignment = center_align
            ws[f'A{summary_row}'].fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

            wb.save(file_path)
            messagebox.showinfo("Éxito", f"Excel exportado correctamente:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exportando a Excel: {str(e)}")

    def _export_pdf(self):
        """Exporta datos de clientes a PDF"""
        try:
            clientes = self.controller.get_all()
            if not clientes:
                messagebox.showerror("Error", "No hay datos para exportar")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Guardar PDF como",
                initialfile="clientes.pdf"
            )

            if not file_path:
                return

            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1,
                textColor=colors.darkblue
            )

            elements = []

            # Título
            title = Paragraph("REPORTE DE CLIENTES", title_style)
            elements.append(title)

            # Fecha
            fecha = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal'])
            elements.append(fecha)
            elements.append(Spacer(1, 20))

            # Tabla
            table_data = [['ID', 'Nombre', 'Apellido', 'Documento', 'Teléfono', 'Correo']]

            for cliente in clientes:
                table_data.append([
                    safe_str(cliente.get('ID_CLIENTE', '')),
                    safe_str(cliente.get('NOMBRE', '')),
                    safe_str(cliente.get('APELLIDO', '')),
                    safe_str(cliente.get('DOCUMENTO_IDENTIDAD', '')),
                    safe_str(cliente.get('TELEFONO', '')),
                    safe_str(cliente.get('CORREO', ''))
                ])

            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            # Alternar colores
            for i in range(1, len(table_data)):
                bg_color = colors.HexColor('#f8f8f8') if i % 2 == 0 else colors.white
                table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), bg_color)]))

            elements.append(table)
            elements.append(Spacer(1, 20))

            # Resumen
            total_clientes = len(clientes)
            resumen = Paragraph(f"Total de clientes registrados: {total_clientes}", styles['Normal'])
            elements.append(resumen)

            doc.build(elements)
            messagebox.showinfo("Éxito", f"PDF exportado correctamente:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exportando a PDF: {str(e)}")

    def _change_theme(self):
        """Cambia el tema de la aplicación"""
        try:
            self.dark_theme = not self.dark_theme
            self._apply_theme()
            messagebox.showinfo("Tema Cambiado", f"Tema {'oscuro' if self.dark_theme else 'claro'} aplicado")
        except Exception as e:
            messagebox.showerror("Error", f"Error cambiando tema: {str(e)}")

    def _apply_theme(self):
        """Aplica el tema actual a todos los widgets"""
        try:
            if self.dark_theme:
                bg_color = "#2b2b2b"
                fg_color = "#ffffff"
                entry_bg = "#3c3c3c"
                entry_fg = "#ffffff"
                button_bg = "#4CAF50"
                tree_bg = "#3c3c3c"
                tree_fg = "#ffffff"
                tree_heading_bg = "#2E8B57"
            else:
                bg_color = "#f0f0f0"
                fg_color = "#000000"
                entry_bg = "#ffffff"
                entry_fg = "#000000"
                button_bg = "#4CAF50"
                tree_bg = "#ffffff"
                tree_fg = "#000000"
                tree_heading_bg = "#4CAF50"

            # Aplicar colores a todos los contenedores y widgets
            self._apply_theme_to_container(self.main_frame, bg_color, fg_color)

            # Configurar Treeview
            style = ttk.Style()
            style.configure("Treeview",
                            background=tree_bg,
                            foreground=tree_fg,
                            fieldbackground=tree_bg)
            style.configure("Treeview.Heading",
                            background=tree_heading_bg,
                            foreground="white")

        except Exception as e:
            print(f"Error aplicando tema: {e}")

    def _apply_theme_to_container(self, container, bg_color, fg_color):
        """Aplica el tema recursivamente a un contenedor y sus hijos"""
        try:
            container.configure(bg=bg_color)

            for child in container.winfo_children():
                widget_type = child.winfo_class()

                if widget_type in ('Frame', 'Labelframe', 'TFrame'):
                    self._apply_theme_to_container(child, bg_color, fg_color)
                elif widget_type == 'Label':
                    child.configure(bg=bg_color, fg=fg_color)
                elif widget_type == 'Entry':
                    child.configure(bg="#ffffff" if not self.dark_theme else "#3c3c3c",
                                    fg=fg_color, insertbackground=fg_color)
                elif widget_type == 'Button':
                    # Mantener colores originales de botones
                    pass
                elif widget_type == 'Text':
                    child.configure(bg="#ffffff" if not self.dark_theme else "#3c3c3c",
                                    fg=fg_color, insertbackground=fg_color)

                # Aplicar recursivamente a hijos de este widget
                if hasattr(child, 'winfo_children'):
                    for grandchild in child.winfo_children():
                        self._apply_theme_to_widget(grandchild, bg_color, fg_color)

        except Exception as e:
            print(f"Error aplicando tema al contenedor: {e}")

    def _apply_theme_to_widget(self, widget, bg_color, fg_color):
        """Aplica el tema a un widget individual"""
        try:
            widget_type = widget.winfo_class()

            if widget_type in ('Frame', 'Labelframe', 'TFrame'):
                widget.configure(bg=bg_color)
                for child in widget.winfo_children():
                    self._apply_theme_to_widget(child, bg_color, fg_color)
            elif widget_type == 'Label':
                widget.configure(bg=bg_color, fg=fg_color)
            elif widget_type == 'Entry':
                widget.configure(bg="#ffffff" if not self.dark_theme else "#3c3c3c",
                                 fg=fg_color, insertbackground=fg_color)
            elif widget_type == 'Button':
                # Mantener colores originales de botones
                pass
            elif widget_type == 'Text':
                widget.configure(bg="#ffffff" if not self.dark_theme else "#3c3c3c",
                                 fg=fg_color, insertbackground=fg_color)

        except Exception as e:
            print(f"Error aplicando tema al widget: {e}")

    def _clear_form(self):
        """Limpia el formulario y resetea validaciones"""
        super()._clear_form()
        self.focus_field('id_cliente')
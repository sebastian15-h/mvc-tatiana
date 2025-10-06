"""
Vista específica para la gestión de hoteles
"""
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
from views.base_view import BaseView
from utils.helpers import UIHelpers, safe_str
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.exceptions import EntityNotFoundError, EntityInUseError


class HotelesView(BaseView):
    """Vista específica para hoteles"""

    def __init__(self, parent_frame, controller):
        self.form_title = "GESTIÓN DE HOTELES"
        self.entity_name = "HOTELES"
        self.dark_theme = True
        super().__init__(parent_frame, controller)

    def _create_form_fields(self):
        """Crea los campos específicos del formulario de hoteles"""
        # Campo ID_HOTEL
        self.create_form_field(0, "ID_HOTEL:", "id_hotel")

        # Campo NOMBRE_HOTEL
        self.create_form_field(1, "NOMBRE_HOTEL:", "nombre_hotel")

        # Campo CATEGORIA
        self.create_form_field(2, "CATEGORIA:", "categoria")

        # Campo DIRECCION
        self.create_form_field(3, "DIRECCION:", "direccion")

        # Campo TELEFONO
        self.create_form_field(4, "TELEFONO:", "telefono")

        # Campo CORREO
        self.create_form_field(5, "CORREO:", "correo")

        # Campo AÑO_INAUGURACION
        self.create_form_field(6, "AÑO_INAUGURACION:", "año_inauguracion")

        # Campo HABITANTES
        self.create_form_field(7, "HABITANTES:", "habitantes")

        # Campo SERVICIOS
        self.create_form_field(8, "SERVICIOS:", "servicios")

        # Campo CHECKIN
        self.create_form_field(9, "CHECK-IN HORARIOS:", "checkin")

        # Campo CHECKOUT
        self.create_form_field(10, "CHECK-OUT HORARIOS:", "checkout")

        # Campo GERENTE
        self.create_form_field(11, "GERENTE:", "gerente")

    def _create_buttons(self):
        """Crea los botones de acción - Sobrescribe el método de BaseView"""
        # Frame para botones principales
        main_button_frame = tk.Frame(self.button_frame)
        main_button_frame.pack(pady=10)

        # Botón Guardar (Crear nuevo)
        btn_save = tk.Button(
            main_button_frame,
            text="Guardar",
            font=("Arial", 10, "bold"),
            bg="#4CAF50",
            fg="white",
            width=10,
            command=self._on_save
        )
        btn_save.pack(side=tk.LEFT, padx=3)

        # Botón Actualizar
        btn_update = tk.Button(
            main_button_frame,
            text="Actualizar",
            font=("Arial", 10, "bold"),
            bg="#2196F3",
            fg="white",
            width=10,
            command=self._on_update
        )
        btn_update.pack(side=tk.LEFT, padx=3)

        # Botón Eliminar
        btn_delete = tk.Button(
            main_button_frame,
            text="Eliminar",
            font=("Arial", 10, "bold"),
            bg="#f44336",
            fg="white",
            width=10,
            command=self._on_delete
        )
        btn_delete.pack(side=tk.LEFT, padx=3)

        # Botón Limpiar
        btn_clear = tk.Button(
            main_button_frame,
            text="Limpiar",
            font=("Arial", 10, "bold"),
            bg="#FF9800",
            fg="white",
            width=10,
            command=self._clear_form
        )
        btn_clear.pack(side=tk.LEFT, padx=3)

        # Frame para búsqueda
        search_frame = tk.Frame(self.button_frame)
        search_frame.pack(pady=5)

        # Botón Buscar
        btn_search = tk.Button(
            search_frame,
            text="Buscar por ID",
            font=("Arial", 10),
            bg="#9C27B0",
            fg="white",
            width=12,
            command=self._on_search
        )
        btn_search.pack(side=tk.LEFT, padx=3)

        # Luego añadir los botones adicionales
        self._create_additional_buttons()

    def _create_additional_buttons(self):
        """Crea botones adicionales específicos para hoteles"""
        # Frame para botones adicionales
        additional_button_frame = tk.Frame(self.button_frame)
        additional_button_frame.pack(pady=10)

        # Botón Exportar Excel
        btn_export_excel = tk.Button(
            additional_button_frame,
            text="Exportar Excel",
            font=("Arial", 10),
            bg="#4CAF50",
            fg="white",
            width=12,
            command=self._export_excel
        )
        btn_export_excel.pack(side=tk.LEFT, padx=3)

        # Botón Exportar PDF
        btn_export_pdf = tk.Button(
            additional_button_frame,
            text="Exportar PDF",
            font=("Arial", 10),
            bg="#2196F3",
            fg="white",
            width=12,
            command=self._export_pdf
        )
        btn_export_pdf.pack(side=tk.LEFT, padx=3)

        # Botón Cambiar Tema
        btn_change_theme = tk.Button(
            additional_button_frame,
            text="Cambiar Tema",
            font=("Arial", 10),
            bg="#9E9E9E",
            fg="white",
            width=12,
            command=self._change_theme
        )
        btn_change_theme.pack(side=tk.LEFT, padx=3)

    def _create_treeview(self, parent):
        """Crea el TreeView específico para hoteles"""
        # Frame para TreeView y scrollbar
        tree_frame = tk.Frame(parent)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Definir columnas
        columns = ('ID_HOTEL', 'NOMBRE_HOTEL', 'CATEGORIA', 'DIRECCION',
                  'TELEFONO', 'CORREO', 'AÑO_INAUGURACION', 'HABITANTES',
                  'SERVICIOS', 'CHECKIN', 'CHECKOUT', 'GERENTE')

        # Crear TreeView
        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            height=20
        )

        # Configurar encabezados
        self.tree.heading('ID_HOTEL', text='ID')
        self.tree.heading('NOMBRE_HOTEL', text='NOMBRE')
        self.tree.heading('CATEGORIA', text='CATEGORÍA')
        self.tree.heading('DIRECCION', text='DIRECCIÓN')
        self.tree.heading('TELEFONO', text='TELÉFONO')
        self.tree.heading('CORREO', text='CORREO')
        self.tree.heading('AÑO_INAUGURACION', text='INAUGURACIÓN')
        self.tree.heading('HABITANTES', text='HABITANTES')
        self.tree.heading('SERVICIOS', text='SERVICIOS')
        self.tree.heading('CHECKIN', text='CHECK-IN')
        self.tree.heading('CHECKOUT', text='CHECK-OUT')
        self.tree.heading('GERENTE', text='GERENTE')

        # Configurar anchos de columnas
        self.tree.column('ID_HOTEL', width=50, anchor='center')
        self.tree.column('NOMBRE_HOTEL', width=120, anchor='w')
        self.tree.column('CATEGORIA', width=80, anchor='center')
        self.tree.column('DIRECCION', width=120, anchor='w')
        self.tree.column('TELEFONO', width=100, anchor='center')
        self.tree.column('CORREO', width=120, anchor='w')
        self.tree.column('AÑO_INAUGURACION', width=100, anchor='center')
        self.tree.column('HABITANTES', width=80, anchor='center')
        self.tree.column('SERVICIOS', width=120, anchor='w')
        self.tree.column('CHECKIN', width=80, anchor='center')
        self.tree.column('CHECKOUT', width=80, anchor='center')
        self.tree.column('GERENTE', width=100, anchor='w')

        # Configurar estilos iniciales
        self._configure_treeview_styles()

        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Layout
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Evento de selección - ¡ESTO ES CLAVE!
        self.tree.bind('<<TreeviewSelect>>', self._on_tree_select)

    def _on_tree_select(self, event):
        """Maneja la selección de un item en el treeview"""
        try:
            selected_items = self.tree.selection()
            if not selected_items:
                return

            selected_item = selected_items[0]
            values = self.tree.item(selected_item, 'values')

            if values and len(values) >= 12:
                data = {
                    'ID_HOTEL': values[0],
                    'NOMBRE_HOTEL': values[1],
                    'CATEGORIA': values[2],
                    'DIRECCION': values[3],
                    'TELEFONO': values[4],
                    'CORREO': values[5],
                    'AÑO_INAUGURACION': values[6],
                    'HABITANTES': values[7],
                    'SERVICIOS': values[8],
                    'CHECKIN': values[9],
                    'CHECKOUT': values[10],
                    'GERENTE': values[11]
                }
                print(f"Datos seleccionados del treeview: {data}")
                self._populate_form(data)

        except Exception as e:
            print(f"Error en selección de treeview: {e}")

    def _configure_treeview_styles(self, bg_color=None, fg_color=None, heading_bg=None, odd_bg=None, even_bg=None):
        """Configura los estilos del TreeView"""
        try:
            style = ttk.Style()
            style.theme_use("default")

            # Usar valores por defecto si no se proporcionan
            if bg_color is None:
                bg_color = "#2b2b2b" if self.dark_theme else "white"
            if fg_color is None:
                fg_color = "white" if self.dark_theme else "black"
            if heading_bg is None:
                heading_bg = "#2ecc71" if self.dark_theme else "#4CAF50"
            if odd_bg is None:
                odd_bg = "#1e1e1e" if self.dark_theme else "#f9f9f9"
            if even_bg is None:
                even_bg = "#2e2e2e" if self.dark_theme else "#e9e9e9"

            # Configurar TreeView
            style.configure("Treeview",
                            background=bg_color,
                            foreground=fg_color,
                            rowheight=25,
                            fieldbackground=bg_color,
                            font=('Arial', 11))

            # Configurar encabezados
            style.configure("Treeview.Heading",
                            background=heading_bg,
                            foreground="white",
                            font=('Arial', 12, 'bold'))

            # Colores alternados para filas
            self.tree.tag_configure('oddrow', background=odd_bg)
            self.tree.tag_configure('evenrow', background=even_bg)

        except Exception as e:
            print(f"Error configurando estilos: {e}")

    def _get_form_data(self):
        """Obtiene los datos del formulario de hoteles"""
        form_data = {
            'ID_HOTEL': self.get_field_value('id_hotel'),
            'NOMBRE_HOTEL': self.get_field_value('nombre_hotel'),
            'CATEGORIA': self.get_field_value('categoria'),
            'DIRECCION': self.get_field_value('direccion'),
            'TELEFONO': self.get_field_value('telefono'),
            'CORREO': self.get_field_value('correo'),
            'AÑO_INAUGURACION': self.get_field_value('año_inauguracion'),
            'HABITANTES': self.get_field_value('habitantes'),
            'SERVICIOS': self.get_field_value('servicios'),
            'CHECKIN': self.get_field_value('checkin'),
            'CHECKOUT': self.get_field_value('checkout'),
            'GERENTE': self.get_field_value('gerente')
        }
        print(f"Datos obtenidos del formulario: {form_data}")
        return form_data

    def _populate_form(self, data):
        """Puebla el formulario con datos de un hotel"""
        try:
            if not data:
                return

            print(f"Poblando formulario con datos: {data}")

            # Limpiar formulario primero
            self._clear_form()

            # Poblar campos con los datos
            self.set_field_value('id_hotel', data.get('ID_HOTEL', ''))
            self.set_field_value('nombre_hotel', data.get('NOMBRE_HOTEL', ''))
            self.set_field_value('categoria', data.get('CATEGORIA', ''))
            self.set_field_value('direccion', data.get('DIRECCION', ''))
            self.set_field_value('telefono', data.get('TELEFONO', ''))
            self.set_field_value('correo', data.get('CORREO', ''))
            self.set_field_value('año_inauguracion', data.get('AÑO_INAUGURACION', ''))
            self.set_field_value('habitantes', data.get('HABITANTES', ''))
            self.set_field_value('servicios', data.get('SERVICIOS', ''))
            self.set_field_value('checkin', data.get('CHECKIN', ''))
            self.set_field_value('checkout', data.get('CHECKOUT', ''))
            self.set_field_value('gerente', data.get('GERENTE', ''))

            print("Formulario poblado exitosamente")

        except Exception as e:
            print(f"Error poblando formulario: {e}")

    def _get_entity_id_from_form(self):
        """Obtiene el ID del hotel desde el formulario"""
        hotel_id = self.get_field_value('id_hotel')
        try:
            return int(hotel_id) if hotel_id and hotel_id.strip() else None
        except ValueError:
            return None

    def _refresh_list(self):
        """Actualiza la lista de hoteles con colores alternados"""
        try:
            # Limpiar TreeView completamente
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Obtener datos actualizados del controlador
            hoteles = self.controller.get_all()
            print(f"Refrescando lista con {len(hoteles)} hoteles")

            # Insertar datos actualizados
            for index, hotel in enumerate(hoteles):
                values = (
                    safe_str(hotel.get('ID_HOTEL', '')),
                    safe_str(hotel.get('NOMBRE_HOTEL', '')),
                    safe_str(hotel.get('CATEGORIA', '')),
                    safe_str(hotel.get('DIRECCION', '')),
                    safe_str(hotel.get('TELEFONO', '')),
                    safe_str(hotel.get('CORREO', '')),
                    safe_str(hotel.get('AÑO_INAUGURACION', '')),
                    safe_str(hotel.get('HABITANTES', '')),
                    safe_str(hotel.get('SERVICIOS', '')),
                    safe_str(hotel.get('CHECKIN', '')),
                    safe_str(hotel.get('CHECKOUT', '')),
                    safe_str(hotel.get('GERENTE', ''))
                )
                tag = 'evenrow' if index % 2 == 0 else 'oddrow'
                self.tree.insert('', 'end', values=values, tags=(tag,))

            print("Lista refrescada exitosamente")

        except Exception as e:
            print(f"Error en _refresh_list: {e}")
            import traceback
            traceback.print_exc()

    def _on_save(self):
        """Maneja el guardado de hoteles"""
        try:
            form_data = self._get_form_data()
            print("=== INICIANDO GUARDADO ===")
            print("Datos del formulario (GUARDAR):", form_data)

            # Validar campos obligatorios
            if not form_data.get('NOMBRE_HOTEL') or not form_data.get('NOMBRE_HOTEL').strip():
                messagebox.showerror("Error", "Debe ingresar el nombre del hotel")
                return

            if not form_data.get('DIRECCION') or not form_data.get('DIRECCION').strip():
                messagebox.showerror("Error", "Debe ingresar la dirección del hotel")
                return

            # Crear nuevo hotel
            print("Llamando a controller.create...")
            new_id = self.controller.create(form_data)

            if new_id:
                messagebox.showinfo("Éxito", f"Hotel guardado correctamente con ID: {new_id}")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "Error al guardar hotel - No se retornó ID")

        except ValueError as ve:
            messagebox.showerror("Error de validación", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar hotel: {str(e)}")
            print(f"ERROR DETALLADO en _on_save: {e}")
            import traceback
            traceback.print_exc()

    def _on_update(self):
        """Maneja la actualización de hoteles"""
        try:
            hotel_id = self._get_entity_id_from_form()
            print("=== INICIANDO ACTUALIZACIÓN ===")
            print(f"ID para actualizar: {hotel_id}")

            if not hotel_id:
                messagebox.showerror("Error", "Seleccione un hotel para actualizar")
                return

            form_data = self._get_form_data()
            print("Datos del formulario (ACTUALIZAR):", form_data)

            # Validar campos obligatorios
            if not form_data.get('NOMBRE_HOTEL') or not form_data.get('NOMBRE_HOTEL').strip():
                messagebox.showerror("Error", "Debe ingresar el nombre del hotel")
                return

            if not form_data.get('DIRECCION') or not form_data.get('DIRECCION').strip():
                messagebox.showerror("Error", "Debe ingresar la dirección del hotel")
                return

            # Confirmación
            confirmar = messagebox.askyesno("Confirmación", f"¿Desea actualizar el hotel con ID {hotel_id}?")
            if not confirmar:
                return

            print("Llamando a controller.update...")
            # Actualizar hotel
            success = self.controller.update(hotel_id, form_data)

            if success:
                messagebox.showinfo("Éxito", f"Hotel con ID {hotel_id} actualizado correctamente")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "No se pudo actualizar el hotel")

        except ValueError as ve:
            messagebox.showerror("Error de validación", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Error actualizando hotel: {str(e)}")
            print(f"ERROR DETALLADO en _on_update: {e}")
            import traceback
            traceback.print_exc()

    def _on_delete(self):
        """Elimina un hotel"""
        try:
            hotel_id = self._get_entity_id_from_form()
            print("=== INICIANDO ELIMINACIÓN ===")
            print(f"ID para eliminar: {hotel_id}")

            if not hotel_id:
                messagebox.showwarning("Advertencia", "Seleccione un hotel para eliminar")
                return

            # Confirmación
            respuesta = messagebox.askyesno("Confirmar eliminación",
                                          f"¿Estás seguro de eliminar el hotel con ID {hotel_id}?")
            if not respuesta:
                messagebox.showinfo("Cancelado", "Eliminación cancelada")
                return

            print("Llamando a controller.delete...")
            success = self.controller.delete(hotel_id)

            if success:
                messagebox.showinfo("Éxito", f"Hotel con ID {hotel_id} eliminado correctamente")
                self._clear_form()
                self._refresh_list()
            else:
                messagebox.showerror("Error", "No se pudo eliminar el hotel")

        except EntityNotFoundError as e:
            messagebox.showerror("No encontrado", f"No se encontró el hotel con ID {hotel_id}")
            self._clear_form()
            self._refresh_list()
        except EntityInUseError as e:
            messagebox.showerror(
                "No se puede eliminar",
                f"No se puede eliminar el hotel porque {e.reason}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar el hotel: {str(e)}")
            print(f"ERROR DETALLADO en _on_delete: {e}")
            import traceback
            traceback.print_exc()

    def _on_search(self):
        """Busca un hotel por ID"""
        hotel_id = self.get_field_value('id_hotel')

        try:
            print("=== INICIANDO BÚSQUEDA ===")
            print(f"ID a buscar: {hotel_id}")

            if not hotel_id or not hotel_id.strip():
                messagebox.showerror("Error", "Ingrese un ID de hotel para buscar")
                return

            # Validar que sea un número válido
            entity_id = int(hotel_id.strip())

            if entity_id <= 0:
                messagebox.showerror("Error", "El ID del hotel debe ser un número positivo")
                return

            print("Llamando a controller.get_by_id...")
            # Llamar al controlador para buscar
            data = self.controller.get_by_id(entity_id)

            if data:
                self._populate_form(data)
                messagebox.showinfo("Éxito", "Hotel encontrado y cargado en el formulario")
            else:
                messagebox.showerror("No encontrado", f"No se encontró ningún hotel con ID: {entity_id}")

        except ValueError:
            messagebox.showerror("Error", "El ID del hotel debe ser un número válido")
        except Exception as e:
            messagebox.showerror("Error", f"Error en búsqueda: {str(e)}")
            print(f"ERROR DETALLADO en _on_search: {e}")

    def _change_theme(self):
        """Cambia el tema de la aplicación"""
        try:
            self.dark_theme = not self.dark_theme
            self._apply_theme()
            messagebox.showinfo("Tema Cambiado", f"Tema {'oscuro' if self.dark_theme else 'claro'} aplicado")
        except Exception as e:
            messagebox.showerror("Error", f"Error cambiando tema: {str(e)}")

    def _apply_theme(self):
        """Aplica el tema actual a todos los elementos de la vista"""
        try:
            if self.dark_theme:
                # Tema oscuro
                bg_color = "#2b2b2b"
                fg_color = "white"
                entry_bg = "#1e1e1e"
                entry_fg = "white"
                button_bg = "#2196F3"
                tree_bg = "#2b2b2b"
                tree_fg = "white"
                heading_bg = "#2ecc71"
                odd_bg = "#1e1e1e"
                even_bg = "#2e2e2e"
            else:
                # Tema claro
                bg_color = "#f0f0f0"
                fg_color = "black"
                entry_bg = "white"
                entry_fg = "black"
                button_bg = "#2196F3"
                tree_bg = "white"
                tree_fg = "black"
                heading_bg = "#4CAF50"
                odd_bg = "#f9f9f9"
                even_bg = "#e9e9e9"

            # Aplicar colores a frames principales
            self.main_frame.configure(bg=bg_color)
            self.form_frame.configure(bg=bg_color)
            self.button_frame.configure(bg=bg_color)

            # Aplicar tema recursivamente
            self._apply_theme_to_container(self.main_frame, bg_color, fg_color, entry_bg, entry_fg, button_bg)
            self._apply_theme_to_container(self.form_frame, bg_color, fg_color, entry_bg, entry_fg, button_bg)
            self._apply_theme_to_container(self.button_frame, bg_color, fg_color, entry_bg, entry_fg, button_bg)

            # Aplicar tema al TreeView
            self._configure_treeview_styles(tree_bg, tree_fg, heading_bg, odd_bg, even_bg)

        except Exception as e:
            print(f"Error aplicando tema: {e}")

    def _apply_theme_to_container(self, container, bg_color, fg_color, entry_bg, entry_fg, button_bg):
        """Aplica el tema recursivamente a un contenedor y todos sus hijos"""
        try:
            # Aplicar al contenedor mismo
            if isinstance(container, (tk.Frame, tk.LabelFrame)):
                container.configure(bg=bg_color)

            # Aplicar a todos los hijos del contenedor
            for widget in container.winfo_children():
                self._apply_theme_to_widget(widget, bg_color, fg_color, entry_bg, entry_fg, button_bg)

                # Si el widget es un contenedor, aplicar recursivamente a sus hijos
                if isinstance(widget, (tk.Frame, tk.LabelFrame)):
                    self._apply_theme_to_container(widget, bg_color, fg_color, entry_bg, entry_fg, button_bg)

        except Exception as e:
            print(f"Error aplicando tema al contenedor: {e}")

    def _apply_theme_to_widget(self, widget, bg_color, fg_color, entry_bg, entry_fg, button_bg):
        """Aplica el tema a un widget específico"""
        try:
            if isinstance(widget, tk.Label):
                widget.configure(bg=bg_color, fg=fg_color)
            elif isinstance(widget, tk.Entry):
                widget.configure(bg=entry_bg, fg=entry_fg, insertbackground=fg_color)
            elif isinstance(widget, tk.Button):
                current_text = widget.cget('text')
                # Mantener colores específicos para botones de acción
                if current_text in ['Guardar', 'Actualizar', 'Eliminar', 'Limpiar', 'Buscar']:
                    pass
                elif current_text == 'Cambiar Tema':
                    widget.configure(bg="#9E9E9E", fg="white")
                elif current_text == 'Exportar Excel':
                    widget.configure(bg="#4CAF50", fg="white")
                elif current_text == 'Exportar PDF':
                    widget.configure(bg="#2196F3", fg="white")
                else:
                    widget.configure(bg=button_bg, fg="white")
            elif isinstance(widget, tk.Frame):
                widget.configure(bg=bg_color)
        except Exception as e:
            pass

    def _export_excel(self):
        """Exporta datos de hoteles a Excel con formato profesional"""
        try:
            hoteles = self.controller.get_all()
            if not hoteles:
                messagebox.showerror("Error", "No hay datos para exportar")
                return

            # Pedir al usuario donde guardar el Excel
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar Excel como"
            )

            if not file_path:
                return  # Usuario canceló

            # Crear libro de trabajo y hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Hoteles"

            # Estilos
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
            normal_font = Font(name='Arial', size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')

            # Título
            ws.merge_cells('A1:L1')
            ws['A1'] = 'REPORTE DE HOTELES'
            ws['A1'].font = Font(name='Arial', size=14, bold=True)
            ws['A1'].alignment = center_align

            # Fecha
            ws.merge_cells('A2:L2')
            ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
            ws['A2'].font = normal_font
            ws['A2'].alignment = center_align

            # Espacio
            ws.append([])

            # Encabezados de columnas
            headers = [
                'ID HOTEL',
                'NOMBRE HOTEL',
                'CATEGORÍA',
                'DIRECCIÓN',
                'TELÉFONO',
                'CORREO',
                'AÑO INAUGURACIÓN',
                'NÚMERO HABITANTES',
                'SERVICIOS DISPONIBLES',
                'HORARIOS CHECK-IN',
                'HORARIOS CHECK-OUT',
                'GERENTE RESPONSABLE'
            ]
            ws.append(headers)

            # Aplicar estilo a los encabezados
            for col in range(1, 13):  # Columnas A a L
                cell = ws.cell(row=4, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            # Datos de los hoteles
            for hotel in hoteles:
                row = [
                    safe_str(hotel.get('ID_HOTEL', '')),
                    safe_str(hotel.get('NOMBRE_HOTEL', '')),
                    safe_str(hotel.get('CATEGORIA', '')),
                    safe_str(hotel.get('DIRECCION', '')),
                    safe_str(hotel.get('TELEFONO', '')),
                    safe_str(hotel.get('CORREO', '')),
                    safe_str(hotel.get('AÑO_INAUGURACION', '')),
                    safe_str(hotel.get('HABITANTES', '')),
                    safe_str(hotel.get('SERVICIOS', '')),
                    safe_str(hotel.get('CHECKIN', '')),
                    safe_str(hotel.get('CHECKOUT', '')),
                    safe_str(hotel.get('GERENTE', ''))
                ]
                ws.append(row)

            # Aplicar estilo a los datos
            for row in range(5, len(hoteles) + 5):
                # Alternar colores de fila
                if row % 2 == 1:
                    fill_color = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
                else:
                    fill_color = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                for col in range(1, 13):
                    cell = ws.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = border
                    cell.fill = fill_color

                    # Alineación específica por columna
                    if col in [1, 3, 7]:  # ID, Categoría, Habitantes - centrados
                        cell.alignment = center_align
                    else:  # Texto - alineado a la izquierda
                        cell.alignment = left_align

            # Ajustar anchos de columna
            column_widths = {
                'A': 10,   # ID
                'B': 25,   # Nombre Hotel
                'C': 12,   # Categoría
                'D': 30,   # Dirección
                'E': 15,   # Teléfono
                'F': 25,   # Correo
                'G': 15,   # Año Inauguración
                'H': 15,   # Habitantes
                'I': 30,   # Servicios
                'J': 15,   # Check-in
                'K': 15,   # Check-out
                'L': 20    # Gerente
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Resumen
            summary_row = len(hoteles) + 6
            ws.merge_cells(f'A{summary_row}:L{summary_row}')
            ws[f'A{summary_row}'] = f'Total de hoteles registrados: {len(hoteles)}'
            ws[f'A{summary_row}'].font = Font(name='Arial', size=11, bold=True)
            ws[f'A{summary_row}'].alignment = center_align
            ws[f'A{summary_row}'].fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

            # Guardar el archivo
            wb.save(file_path)

            messagebox.showinfo("Éxito", f"Excel exportado correctamente:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exportando a Excel: {str(e)}")

    def _export_pdf(self):
        """Exporta datos de hoteles a PDF"""
        try:
            hoteles = self.controller.get_all()
            if not hoteles:
                messagebox.showerror("Error", "No hay datos para exportar")
                return

            # Pedir al usuario donde guardar el PDF
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Guardar PDF como"
            )

            if not file_path:
                return  # Usuario canceló

            # Crear el documento PDF
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )

            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1,  # Centrado
                textColor=colors.darkblue
            )

            # Contenido del PDF
            elements = []

            # Título
            title = Paragraph("REPORTE DE HOTELES", title_style)
            elements.append(title)

            # Fecha de generación
            fecha = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal'])
            elements.append(fecha)
            elements.append(Spacer(1, 20))

            # Preparar datos para la tabla
            table_data = [
                ['ID', 'Nombre Hotel', 'Categoría', 'Dirección', 'Teléfono', 'Correo']
            ]

            for hotel in hoteles:
                table_data.append([
                    safe_str(hotel.get('ID_HOTEL', '')),
                    safe_str(hotel.get('NOMBRE_HOTEL', '')),
                    safe_str(hotel.get('CATEGORIA', '')),
                    safe_str(hotel.get('DIRECCION', '')),
                    safe_str(hotel.get('TELEFONO', '')),
                    safe_str(hotel.get('CORREO', ''))
                ])

            # Crear tabla
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                # Estilo del encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),

                # Estilo de las filas
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),

                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            # Alternar colores de filas
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    bg_color = colors.HexColor('#f8f8f8')
                else:
                    bg_color = colors.white
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), bg_color)
                ]))

            elements.append(table)
            elements.append(Spacer(1, 20))

            # Resumen
            total_hoteles = len(hoteles)
            resumen = Paragraph(f"Total de hoteles registrados: {total_hoteles}", styles['Normal'])
            elements.append(resumen)

            # Generar PDF
            doc.build(elements)

            messagebox.showinfo("Éxito", f"PDF exportado correctamente:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exportando a PDF: {str(e)}")

    def _clear_form(self):
        """Limpia el formulario y resetea validaciones"""
        super()._clear_form()

        # Enfocar el primer campo
        self.focus_field('id_hotel')